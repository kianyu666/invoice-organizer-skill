#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成发票 Excel 统计表（openpyxl）。

包含两个 Sheet：
  - 类型统计：按类别统计数量与金额合计；替票为独立类型，不并入普通发票；
              数量用 COUNTIF、金额合计用 SUMIF 公式动态引用「发票明细」Sheet。
  - 发票明细：逐张列出类型、文件名、日期、金额、人员。

输入明细支持 JSON 或 CSV，字段（中英文均可）：
  filename/文件名, type/类型, date/日期, amount/金额, person/人员, is_ticket/是否替票

用法：
    python build_stats.py <目录路径> <明细.json|csv> <输出.xlsx>
依赖：
    pip install openpyxl
"""
import argparse
import csv
import json
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[错误] 缺少 openpyxl，请先安装：pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def parse_amount(value):
    """将金额字段转为 float，失败返回 None。"""
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(",", "").replace("¥", "").replace("￥", ""))
    except ValueError:
        return None


def norm_type(record):
    """标准化类型：替票统一为独立类型『替票』，其余保留原类型。"""
    type_val = str(record.get("type", "") or "").strip()
    is_ticket = str(record.get("is_ticket", "") or "").strip().lower()
    if is_ticket in ("1", "true", "yes", "是", "y"):
        return "替票"
    if "替票" in type_val:
        return "替票"
    return type_val or "未分类"


def load_records(path):
    """从 JSON 或 CSV 加载明细记录，返回规范化记录列表。"""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".json":
        with open(path, "r", encoding="utf-8") as fp:
            data = json.load(fp)
        if not isinstance(data, list):
            print("[错误] JSON 明细顶层必须是数组", file=sys.stderr)
            sys.exit(1)
        return data
    elif ext in (".csv", ".txt"):
        records = []
        with open(path, "r", encoding="utf-8-sig", newline="") as fp:
            reader = csv.DictReader(fp)
            for row in reader:
                records.append(row)
        return records
    else:
        print(f"[错误] 不支持的明细格式：{path}（仅支持 JSON/CSV）", file=sys.stderr)
        sys.exit(1)


def build_workbook(records):
    """构建含「类型统计」「发票明细」两个 Sheet 的工作簿。"""
    wb = Workbook()

    # 规范化明细记录（金额非法则打印警告并跳过）
    clean = []
    for rec in records:
        amount = parse_amount(rec.get("amount"))
        if amount is None:
            print(f"[警告] 金额字段缺失或非法，跳过该条：{rec}", file=sys.stderr)
            continue
        clean.append({
            "type": norm_type(rec),
            "filename": str(rec.get("filename", "") or ""),
            "date": str(rec.get("date", "") or ""),
            "amount": amount,
            "person": str(rec.get("person", "") or ""),
        })

    # ---- 发票明细 Sheet ----
    detail_ws = wb.active
    detail_ws.title = "发票明细"
    detail_ws.append(["类型", "文件名", "日期", "金额", "人员"])
    for i, r in enumerate(clean, start=2):
        detail_ws.cell(row=i, column=1, value=r["type"])
        detail_ws.cell(row=i, column=2, value=r["filename"])
        detail_ws.cell(row=i, column=3, value=r["date"])
        detail_ws.cell(row=i, column=4, value=r["amount"])
        detail_ws.cell(row=i, column=5, value=r["person"])

    # ---- 类型统计 Sheet ----
    stats_ws = wb.create_sheet("类型统计")
    stats_ws.append(["类型", "数量", "金额合计"])
    type_order = []
    for r in clean:
        if r["type"] not in type_order:
            type_order.append(r["type"])
    for idx, t in enumerate(type_order, start=2):
        stats_ws.cell(row=idx, column=1, value=t)
        # 数量：COUNTIF 动态引用明细 A 列
        stats_ws.cell(row=idx, column=2, value=f"=COUNTIF('发票明细'!$A:$A,$A{idx})")
        # 金额合计：SUMIF 动态引用明细 A 列（类型）与 D 列（金额）
        stats_ws.cell(row=idx, column=3,
                      value=f"=SUMIF('发票明细'!$A:$A,$A{idx},'发票明细'!$D:$D)")

    # ---- 样式（表头加粗、浅色底纹、统一列宽）----
    for ws, ncols in ((detail_ws, 5), (stats_ws, 3)):
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDEBF7")
            cell.alignment = Alignment(horizontal="center")
        for col in range(1, ncols + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

    return wb


def main():
    parser = argparse.ArgumentParser(description="生成发票 Excel 统计表")
    parser.add_argument("target_dir", help="目录路径")
    parser.add_argument("detail_file", help="明细 JSON/CSV 路径")
    parser.add_argument("output_xlsx", help="输出 xlsx 路径")
    args = parser.parse_args()

    if not os.path.isdir(args.target_dir):
        print(f"[错误] 目录不存在：{args.target_dir}", file=sys.stderr)
        sys.exit(1)
    if not os.path.isfile(args.detail_file):
        print(f"[错误] 明细文件不存在：{args.detail_file}", file=sys.stderr)
        sys.exit(1)
    out_dir = os.path.dirname(os.path.abspath(args.output_xlsx))
    if out_dir and not os.path.isdir(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    records = load_records(args.detail_file)
    wb = build_workbook(records)
    wb.save(args.output_xlsx)
    print(f"[完成] 统计表已生成：{args.output_xlsx}")
    print("  包含 Sheet：『发票明细』、『类型统计』")
    print("  提示：『类型统计』的金额合计使用 SUMIF 公式动态引用『发票明细』。")


if __name__ == "__main__":
    main()
